from flask import Flask, render_template, request, send_file
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font
import matplotlib.pyplot as plt
import io
import gc

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        weekly_income = float(request.form['weekly_income'])
        weekly_expenses = float(request.form['weekly_expenses'])
        available_for_debt = weekly_income - weekly_expenses

        if available_for_debt <= 0:
            return "Your income minus expenses must be greater than zero."

        debts = []
        debt_names = request.form.getlist('debt_name')
        balances = request.form.getlist('balance')
        aprs = request.form.getlist('apr')
        min_payments = request.form.getlist('min_payment')

        for debt_name, balance, apr, min_payment in zip(debt_names, balances, aprs, min_payments):
            if debt_name and float(balance) > 0:
                debts.append({
                    'DebtName': debt_name,
                    'Balance': float(balance),
                    'APR%': float(apr) if float(apr) > 0 else 0.0,  # Handle 0% APR
                    'MinPayment': float(min_payment)
                })

        if not debts:
            return "Please enter at least one valid debt."

        debts_df = pd.DataFrame(debts)
        weeks = []
        payment_table = []
        balance_table = []

        # Create an Excel file that will later be deleted by Python's garbage collector forcibly (this is because this is financial info and could be sensitive)
        excel_stream = io.BytesIO()
        wb = Workbook()
        ws_payment = wb.create_sheet(title='Payments')
        ws_balance = wb.create_sheet(title='Balances')

        currency_style = NamedStyle(name='currency_style', number_format='$#,##0.00')
        bold_font = Font(bold=True)

        payment_headers = ['Week']
        for debt in debts:
            payment_headers.append(f'{debt["DebtName"]} Payment')
        ws_payment.append(payment_headers)

        balance_headers = ['Week']
        for debt in debts:
            balance_headers.append(f'{debt["DebtName"]} Balance After Payment')
        ws_balance.append(balance_headers)

        for cell in ws_payment[1]:
            cell.font = bold_font
            cell.style = currency_style

        for cell in ws_balance[1]:
            cell.font = bold_font
            cell.style = currency_style

        week = 0
        while debts_df['Balance'].sum() > 0:
            week += 1
            weeks.append(week)

            remaining_payment = available_for_debt
            payments = {debt['DebtName']: 0 for debt in debts_df.to_dict('records')}
            debts_df_sorted = debts_df.sort_values(by='APR%', ascending=False)

            for index, debt in debts_df_sorted.iterrows():
                debt_name = debt['DebtName']
                balance = debt['Balance']
                min_payment = debt['MinPayment']

                if remaining_payment > 0:
                    payment = min(min_payment, balance, remaining_payment)
                    payments[debt_name] += payment
                    remaining_payment -= payment
                    debts_df.loc[debts_df['DebtName'] == debt_name, 'Balance'] -= payment

            while remaining_payment > 0 and not debts_df_sorted.empty:
                highest_apr_debt = debts_df_sorted.iloc[0]
                debt_name = highest_apr_debt['DebtName']
                balance = highest_apr_debt['Balance']

                if remaining_payment > 0:
                    payment = min(balance, remaining_payment)
                    payments[debt_name] += payment
                    remaining_payment -= payment
                    new_balance = max(0, debts_df.loc[debts_df['DebtName'] == debt_name, 'Balance'].item() - payment)
                    debts_df.loc[debts_df['DebtName'] == debt_name, 'Balance'] = new_balance
                    debts_df_sorted = debts_df_sorted[debts_df_sorted['DebtName'] != debt_name]

                if debts_df['Balance'].sum() <= 0:
                    break

            payment_row = [week]
            balance_row = [week]
            for i in range(len(debts_df)):
                debt = debts_df.iloc[i]
                debt_name = debt['DebtName']
                balance = debt['Balance']
                payment = payments.get(debt_name, 0)
                new_balance = balance
                payment_row.append(payment)
                balance_row.append(new_balance)

            payment_table.append(payment_row)
            balance_table.append(balance_row)
            ws_payment.append(payment_row)
            ws_balance.append(balance_row)

            # Update for monthly APRs
            for debt in debts_df.itertuples():
                if debt._3 > 0:  # APR% > 0
                    monthly_apr = debt._3 / 12 / 100
                    debts_df.loc[debts_df['DebtName'] == debt.DebtName, 'Balance'] += debts_df['Balance'] * monthly_apr

        for row in ws_payment.iter_rows(min_row=2, min_col=2, max_col=len(payment_headers)):
            for cell in row:
                cell.style = currency_style

        for row in ws_balance.iter_rows(min_row=2, min_col=2, max_col=len(balance_headers)):
            for cell in row:
                cell.style = currency_style

        # Create in-memory plot
        plt.figure(figsize=(10, 6))
        plt.plot(weeks, [sum(row[1:]) for row in balance_table], color='red', marker='o', linestyle='-', label='Debt Balance')
        plt.xlabel('Weeks')
        plt.ylabel('Debt Balance')
        plt.title('Debt Balance Over Time')
        plt.legend()
        plt.grid(True)

        chart_stream = io.BytesIO()
        plt.savefig(chart_stream, format='png')
        plt.close()
        chart_stream.seek(0)

        # Add chart image to Excel file
        img = Image(chart_stream)
        img.width = 800  # Adjust size if necessary
        img.height = 400
        ws_chart = wb.create_sheet(title='Debt Balance Over Time')
        ws_chart.add_image(img, 'A1')

        # Save Excel file to in-memory stream
        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)

        # Force garbage collection
        gc.collect()

        return send_file(excel_stream, as_attachment=True, download_name='weekly_payoff_schedule.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return render_template('index.html')

if __name__ == "__main__":
    app.run(debug=True)
